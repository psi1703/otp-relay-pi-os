"""
generate_sample_users.py
Run once to create a sample data/users.xlsx.
Replace with your real data in the same format.

Columns required:
  token  — 3 uppercase letters, unique per person
  name   — full display name
  email  — company email address
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

os.makedirs("data", exist_ok=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Users"

headers = ["token", "name", "email"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h.upper())
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0f172a")
    cell.alignment = Alignment(horizontal="center")

sample_users = [
    ("AHM", "Ahmed Al Mansoori",  "ahmed@company.com"),
    ("SAR", "Sara Johnson",       "sara@company.com"),
    ("MAR", "Maria Garcia",       "maria@company.com"),
    ("JOH", "John Smith",         "john@company.com"),
    ("FAT", "Fatima Al Zaabi",    "fatima@company.com"),
]

for row_idx, user in enumerate(sample_users, 2):
    for col_idx, value in enumerate(user, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 32

wb.save("data/users.xlsx")
print("✅ data/users.xlsx created.")
print("   Replace the sample rows with your real users before deploying.")
