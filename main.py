# OTP Relay Server — main.py
# Stack: FastAPI + Python 3.12 + Exchange SMTP (internal only)
# No external APIs. Runs entirely on your company LAN.
#
# Delivery model: OTP is displayed on-screen via polling. Email is NOT used
# for OTP delivery. SMTP config and /admin/smtp-test are retained for
# diagnostics only.

import os, re, asyncio, logging, smtplib, json, secrets
from collections import deque
from datetime import datetime, timezone
from typing import Optional, Any, Dict, List
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
from fastapi import FastAPI, Request, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from pydantic import BaseModel, Field
import bcrypt

load_dotenv()

app = FastAPI(title="OTP Relay")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # safe — server is LAN-only
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Config ────────────────────────────────────────────────────────────────────
SMS_SECRET_TOKEN = os.getenv("SMS_SECRET_TOKEN", "changeme")

SMTP_HOST        = os.getenv("SMTP_HOST", "mail.company.local")
SMTP_PORT        = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER        = os.getenv("SMTP_USER", "otp-relay@company.com")
SMTP_PASSWORD    = os.getenv("SMTP_PASSWORD", "")
SMTP_USE_TLS     = os.getenv("SMTP_USE_TLS", "true").lower() == "true"
SMTP_AUTH        = os.getenv("SMTP_AUTH", "true").lower() == "true"
FROM_EMAIL       = os.getenv("FROM_EMAIL", SMTP_USER)
FROM_NAME        = os.getenv("FROM_NAME", "OTP Relay")

# How long the active user has to trigger their OTP before being evicted.
# Other users wait until this window expires or OTP is delivered.
CLAIM_EXPIRY_SEC  = int(os.getenv("CLAIM_EXPIRY_SEC", "90"))

# How long the delivered OTP stays visible on-screen before being purged.
OTP_DISPLAY_SEC   = int(os.getenv("OTP_DISPLAY_SEC", "285"))   # 4 min 45 sec

# If two claims arrive within this window, log a concurrent_risk event.
CONCURRENT_RISK_SEC = int(os.getenv("CONCURRENT_RISK_SEC", "30"))

USERS_EXCEL_PATH = os.getenv("USERS_EXCEL_PATH", "data/users.xlsx")
AUDIT_LOG_PATH   = os.getenv("AUDIT_LOG_PATH", "data/audit.log")

# ── State ─────────────────────────────────────────────────────────────────────
# Queue: max depth 1 enforced at claim time. Others wait and poll.
users: dict        = {}
claim_queue: deque = deque()

# Delivered OTPs held in memory only — never written to disk or logs.
# Structure: { token: { "otp": str, "arrived_at": datetime } }
pending_otps: dict = {}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%SZ",
)
logger = logging.getLogger("otp-relay")


# ── Server-backed wizard/admin state ─────────────────────────────────────────
DATA_DIR = Path(os.environ.get("OTP_RELAY_DATA_DIR", "data"))
WIZARD_FILE = DATA_DIR / "wizard_progress.json"
AUTH_FILE = DATA_DIR / "admin_auth.json"
CONFIG_FILE = DATA_DIR / "admin_config.json"
DEFAULT_ADMIN_TOKENS = ["JPR", "AMD", "SCH"]
ADMIN_TTL_SECONDS = 8 * 60 * 60
ADMIN_SESSIONS: Dict[str, float] = {}

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        logger.warning(f"Could not read {path}: {e}")
        return default

def _write_json(path: Path, payload: Any) -> None:
    _ensure_data_dir()
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")

def _wizard_db() -> Dict[str, dict]:
    return _read_json(WIZARD_FILE, {})

def _save_wizard_db(db: Dict[str, dict]) -> None:
    _write_json(WIZARD_FILE, db)

def _auth_db() -> Dict[str, Any]:
    return _read_json(AUTH_FILE, {})

def _save_auth_db(db: Dict[str, Any]) -> None:
    _write_json(AUTH_FILE, db)

def _config_db() -> Dict[str, Any]:
    env_tokens = os.environ.get("ADMIN_TOKENS", "")
    env_default = [t.strip().upper() for t in env_tokens.split(",") if t.strip()] or DEFAULT_ADMIN_TOKENS
    return _read_json(CONFIG_FILE, {"admin_tokens": env_default})

def _save_config_db(db: Dict[str, Any]) -> None:
    _write_json(CONFIG_FILE, db)

def _purge_admin_sessions() -> None:
    now_ts = datetime.now(timezone.utc).timestamp()
    stale = [s for s, ts in ADMIN_SESSIONS.items() if now_ts - ts > ADMIN_TTL_SECONDS]
    for s in stale:
        ADMIN_SESSIONS.pop(s, None)

def _require_admin(session: Optional[str]) -> None:
    _purge_admin_sessions()
    if not session:
        raise HTTPException(status_code=401, detail="Missing admin session")
    ts = ADMIN_SESSIONS.get(session)
    if not ts:
        raise HTTPException(status_code=401, detail="Invalid admin session")
    ADMIN_SESSIONS[session] = datetime.now(timezone.utc).timestamp()

class WizardRecord(BaseModel):
    token: str
    display_name: str = ""
    iits_username: str = ""
    adm_username: str = ""
    completed: List[str] = Field(default_factory=list)
    adminCompleted: List[str] = Field(default_factory=list)
    iits_pw_date: Optional[str] = None
    adm_pw_date: Optional[str] = None
    vpn_date: Optional[str] = None

class CredentialPayload(BaseModel):
    credential: str
    current: Optional[str] = None

class ConfigPayload(BaseModel):
    admin_tokens: List[str]


# ── User loading ──────────────────────────────────────────────────────────────
def load_users_from_excel(path: str) -> int:
    """
    Reads users.xlsx. Expected columns (row 1 = headers):
      token  — 2 or 3 character unique string, e.g. AH or AHM
      name   — display name
      email  — company email address
    Column names are case-insensitive.
    Skipped rows are written to the audit log so IT can fix them.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    raw_headers = [
        str(c.value).strip().lower() if c.value else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    loaded   = 0
    skipped  = 0
    seen_tokens = {}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        row_dict = dict(zip(raw_headers, row))
        token = str(row_dict.get("token", "") or "").strip().upper()
        name  = str(row_dict.get("name",  "") or "").strip()
        email = str(row_dict.get("email", "") or "").strip()

        if len(token) == 0:
            audit("import_skipped", detail=f"Row {row_num}: empty token — name={repr(name)} email={repr(email)}", status="warn")
            skipped += 1; continue

        if not (2 <= len(token) <= 3):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token must be 2 or 3 characters, got {len(token)} ({repr(token)})", status="warn")
            skipped += 1; continue

        if not re.match(r'^[A-Z0-9]+$', token):
            audit("import_skipped", token=token, detail=f"Row {row_num}: token contains invalid characters ({repr(token)}) — only letters and digits allowed", status="warn")
            skipped += 1; continue

        if not email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: missing email address for {repr(name)}", status="warn")
            skipped += 1; continue

        if "@" not in email:
            audit("import_skipped", token=token, detail=f"Row {row_num}: invalid email address {repr(email)}", status="warn")
            skipped += 1; continue

        if token in seen_tokens:
            audit("import_skipped", token=token, detail=f"Row {row_num}: duplicate token — already defined at row {seen_tokens[token]}", status="warn")
            skipped += 1; continue

        seen_tokens[token] = row_num
        users[token] = {"token": token, "name": name, "email": email}
        loaded += 1

    logger.info(f"Loaded {loaded} users from {path} ({skipped} rows skipped)")
    if skipped > 0:
        audit("import_complete", detail=f"{loaded} users loaded, {skipped} rows skipped — check import_skipped entries above", status="warn")
    else:
        audit("import_complete", detail=f"{loaded} users loaded, no issues")
    return loaded


# ── Audit log ─────────────────────────────────────────────────────────────────
def audit(event: str, token: Optional[str] = None, detail: str = "", status: str = "info"):
    entry = {
        "ts":     datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "event":  event,
        "token":  token or "",
        "detail": detail,
        "status": status,
    }
    try:
        Path(AUDIT_LOG_PATH).parent.mkdir(parents=True, exist_ok=True)
        with open(AUDIT_LOG_PATH, "a") as f:
            f.write(json.dumps(entry) + "\n")
    except Exception as e:
        logger.warning(f"Could not write audit log: {e}")
    level = {"info": logging.INFO, "warn": logging.WARNING, "error": logging.ERROR}.get(status, logging.INFO)
    logger.log(level, f"[{event}] token={token or '—'}  {detail}")


def read_audit_log(limit: int = 200) -> list:
    try:
        lines = Path(AUDIT_LOG_PATH).read_text().strip().splitlines()
        entries = [json.loads(l) for l in lines if l.strip()]
        return list(reversed(entries))[:limit]
    except FileNotFoundError:
        return []
    except Exception as e:
        logger.warning(f"Could not read audit log: {e}")
        return []


# ── Queue and OTP state helpers ───────────────────────────────────────────────
def purge_expired():
    """Evict the front-of-queue claim if it has exceeded CLAIM_EXPIRY_SEC."""
    now = datetime.utcnow()
    while claim_queue:
        age = (now - claim_queue[0]["claimed_at"]).total_seconds()
        if age > CLAIM_EXPIRY_SEC:
            expired = claim_queue.popleft()
            audit("claim_expired", expired["token"],
                  f"No OTP arrived within {CLAIM_EXPIRY_SEC}s — evicted from slot 1", "warn")
        else:
            break


def purge_stale_otps():
    """Remove delivered OTPs that have exceeded OTP_DISPLAY_SEC."""
    now = datetime.utcnow()
    stale = [
        tok for tok, v in pending_otps.items()
        if (now - v["arrived_at"]).total_seconds() > OTP_DISPLAY_SEC
    ]
    for tok in stale:
        del pending_otps[tok]
        audit("otp_display_expired", tok, f"OTP display window closed after {OTP_DISPLAY_SEC}s")


def extract_otp(text: str) -> str:
    match = re.search(r'\b\d{4,8}\b', text)
    return match.group() if match else "—"


# ── Email (diagnostics only — not used for OTP delivery) ─────────────────────
def send_email(to_email: str, name: str, subject: str, html: str):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = f"{FROM_NAME} <{FROM_EMAIL}>"
    msg["To"]      = to_email
    msg.attach(MIMEText(html, "html"))

    if SMTP_USE_TLS:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.ehlo()
        server.starttls()
    else:
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)

    if SMTP_AUTH:
        server.login(SMTP_USER, SMTP_PASSWORD)

    server.sendmail(FROM_EMAIL, to_email, msg.as_string())
    server.quit()


# ── Background task ───────────────────────────────────────────────────────────
async def background_purge():
    """Runs every 15 seconds to expire stale queue entries and OTP display windows."""
    while True:
        await asyncio.sleep(15)
        purge_expired()
        purge_stale_otps()


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    if os.path.exists(USERS_EXCEL_PATH):
        count = load_users_from_excel(USERS_EXCEL_PATH)
        audit("server_start", detail=f"{count} users loaded")
    else:
        logger.warning(f"users.xlsx not found at {USERS_EXCEL_PATH}")
        audit("server_start", detail="No users.xlsx — POST /admin/reload-users after adding it", status="warn")
    asyncio.create_task(background_purge())


@app.post("/claim-otp")
async def claim_otp(request: Request):
    data  = await request.json()
    token = str(data.get("token", "")).strip().upper()

    if token not in users:
        audit("claim_rejected", token, "Unknown token", "error")
        raise HTTPException(status_code=404, detail="Token not recognised. Check with your IT department.")

    purge_expired()
    purge_stale_otps()

    # Already queued — return current status without re-queuing
    for i, claim in enumerate(claim_queue):
        if claim["token"] == token:
            age = (datetime.utcnow() - claim["claimed_at"]).total_seconds()
            remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
            audit("claim_duplicate", token, f"Already at position {i+1}", "warn")
            return {
                "status":     "already_queued",
                "position":   i + 1,
                "expires_in": remaining,
                "queue_depth": len(claim_queue),
            }

    # Already has a delivered OTP waiting on-screen
    if token in pending_otps:
        age = (datetime.utcnow() - pending_otps[token]["arrived_at"]).total_seconds()
        remaining = max(0, int(OTP_DISPLAY_SEC - age))
        return {"status": "otp_ready", "expires_in": remaining}

    # Queue depth = 1 enforced: only one active user at a time.
    # Others are allowed to join the queue and wait — they are NOT allowed to
    # trigger their OTP on the platform until they reach position 1.
    now = datetime.utcnow()

    # Concurrent risk detection: warn if a second claim arrives close behind
    # the current front-of-queue (they could race to trigger OTPs).
    if claim_queue:
        front_age = (now - claim_queue[0]["claimed_at"]).total_seconds()
        if front_age < CONCURRENT_RISK_SEC:
            audit("concurrent_risk", token,
                  f"New claim while {claim_queue[0]['token']} has been active for only {int(front_age)}s",
                  "warn")

    claim_queue.append({
        "token":      token,
        "name":       users[token]["name"],
        "email":      users[token]["email"],
        "claimed_at": now,
    })

    position   = len(claim_queue)
    queue_depth = len(claim_queue)

    # Worst-case wait: each person ahead of them gets the full CLAIM_EXPIRY_SEC.
    # Position 1 = active now, position 2 = up to 1×90s, etc.
    wait_estimate = max(0, (position - 1) * CLAIM_EXPIRY_SEC)

    audit("claim_queued", token, f"Queue position {position} of {queue_depth}")
    return {
        "status":        "queued",
        "position":      position,
        "name":          users[token]["name"],
        "expires_in":    CLAIM_EXPIRY_SEC,
        "queue_depth":   queue_depth,
        "wait_estimate": wait_estimate,   # seconds, worst case
    }


@app.get("/claim-status/{token}")
async def claim_status(token: str):
    token = token.upper()

    purge_expired()
    purge_stale_otps()

    # OTP is ready and waiting on-screen
    if token in pending_otps:
        age = (datetime.utcnow() - pending_otps[token]["arrived_at"]).total_seconds()
        remaining = max(0, int(OTP_DISPLAY_SEC - age))
        return {
            "status":     "delivered",
            "otp":        pending_otps[token]["otp"],
            "expires_in": remaining,
        }

    # Still in the claim queue
    for i, claim in enumerate(claim_queue):
        if claim["token"] == token:
            age       = (datetime.utcnow() - claim["claimed_at"]).total_seconds()
            remaining = max(0, int(CLAIM_EXPIRY_SEC - age))
            # Worst-case wait for users behind position 1
            wait_estimate = max(0, i * CLAIM_EXPIRY_SEC)
            return {
                "status":        "waiting",
                "position":      i + 1,
                "expires_in":    remaining,
                "queue_depth":   len(claim_queue),
                "wait_estimate": wait_estimate,
            }

    # Not in queue, not delivered — check log for recent terminal events
    for e in read_audit_log(500):
        if e.get("token") == token:
            if e["event"] in ("otp_delivered", "otp_display_expired"):
                return {"status": "done"}
            if e["event"] == "claim_expired":
                return {"status": "idle_expired"}
            break

    return {"status": "unknown"}


@app.delete("/claim-otp/{token}")
async def cancel_claim(token: str):
    """
    Discard a delivered OTP and re-queue the user (Retry / Send again flow).
    Also used when user explicitly abandons their slot.
    """
    token = token.upper()

    if token in pending_otps:
        del pending_otps[token]
        audit("otp_discarded", token, "User requested retry — OTP discarded from memory")

    # Remove from queue if present (e.g. user changed their mind while waiting)
    global claim_queue
    before = len(claim_queue)
    claim_queue = deque(c for c in claim_queue if c["token"] != token)
    if len(claim_queue) < before:
        audit("claim_cancelled", token, "Removed from queue by user")

    return {"status": "ok"}


@app.post("/sms-received")
async def sms_received(request: Request):
    if request.headers.get("X-Secret-Token", "") != SMS_SECRET_TOKEN:
        audit("sms_rejected", detail="Wrong secret token", status="error")
        raise HTTPException(status_code=401)

    data     = await request.json()
    sms_body = str(data.get("body", "")).strip()
    audit("sms_received", detail=f"SMS arrived ({len(sms_body)} chars)")

    purge_expired()
    purge_stale_otps()

    if not claim_queue:
        # Brief wait to absorb a race-condition claim that's in-flight
        await asyncio.sleep(4)
        purge_expired()
        if not claim_queue:
            audit("sms_unmatched", detail="No claimant in queue — SMS discarded", status="warn")
            return {"status": "no_claimant"}

    recipient = claim_queue.popleft()
    otp       = extract_otp(sms_body)

    # Store OTP in memory only — never logged, never written to disk.
    pending_otps[recipient["token"]] = {
        "otp":        otp,
        "arrived_at": datetime.utcnow(),
    }

    # Audit record: token and timestamp only, OTP value deliberately omitted.
    audit("otp_delivered", recipient["token"],
          f"OTP ready for display — queue unblocked")

    return {"status": "delivered", "recipient": recipient["name"]}


@app.get("/admin/log")
async def get_log(limit: int = 200):
    entries = read_audit_log(limit)
    return {"entries": entries, "total": len(entries)}


@app.get("/admin/queue")
async def get_queue():
    now = datetime.utcnow()
    return {"queue": [{
        "token":      c["token"],
        "name":       c["name"],
        "email":      c["email"],
        "claimed_at": c["claimed_at"].strftime("%Y-%m-%dT%H:%M:%SZ"),
        "expires_in": max(0, int(CLAIM_EXPIRY_SEC - (now - c["claimed_at"]).total_seconds())),
        "position":   i + 1,
    } for i, c in enumerate(claim_queue)]}


@app.get("/admin/users")
async def list_users():
    return {"count": len(users),
            "users": [{"token": u["token"], "name": u["name"], "email": u["email"]}
                      for u in users.values()]}


@app.post("/admin/reload-users")
async def reload_users():
    if not os.path.exists(USERS_EXCEL_PATH):
        raise HTTPException(status_code=404, detail=f"Not found: {USERS_EXCEL_PATH}")
    users.clear()
    count = load_users_from_excel(USERS_EXCEL_PATH)
    audit("users_reloaded", detail=f"{count} users loaded")
    return {"status": "ok", "users_loaded": count}


@app.get("/admin/smtp-test")
async def smtp_test():
    """Sends a test email to the relay account — use to verify Exchange connectivity."""
    html = """<div style="font-family:Arial,sans-serif;padding:24px">
      <p>OTP Relay SMTP test — if you can read this, Exchange is working. 🎉</p>
    </div>"""
    try:
        send_email(FROM_EMAIL, "OTP Relay", "OTP Relay — SMTP connectivity test", html)
        return {"status": "ok", "sent_to": FROM_EMAIL}
    except Exception as e:
        return {"status": "error", "error": str(e)}




# ── Wizard/admin server-backed endpoints ─────────────────────────────────────
@app.get("/admin/auth/status")
async def admin_auth_status():
    return {"configured": bool(_auth_db().get("password_hash"))}


@app.post("/admin/auth/setup")
async def admin_auth_setup(payload: CredentialPayload):
    cred = (payload.credential or "").strip()
    if len(cred) < 4:
        raise HTTPException(status_code=400, detail="Credential too short")
    db = _auth_db()
    if db.get("password_hash"):
        if not payload.current:
            raise HTTPException(status_code=400, detail="Current credential required")
        if not bcrypt.checkpw(payload.current.encode("utf-8"), db["password_hash"].encode("utf-8")):
            raise HTTPException(status_code=401, detail="Current credential incorrect")
    hashed = bcrypt.hashpw(cred.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    _save_auth_db({"password_hash": hashed, "updated_at": _now_iso()})
    session = secrets.token_urlsafe(24)
    ADMIN_SESSIONS[session] = datetime.now(timezone.utc).timestamp()
    audit("admin_auth_setup", detail="Admin credential configured")
    return {"status": "ok", "session": session}


@app.post("/admin/auth/login")
async def admin_auth_login(payload: CredentialPayload):
    db = _auth_db()
    stored = db.get("password_hash")
    if not stored:
        raise HTTPException(status_code=400, detail="Admin credential not configured")
    if not bcrypt.checkpw((payload.credential or "").encode("utf-8"), stored.encode("utf-8")):
        audit("admin_auth_failed", detail="Incorrect admin credential", status="warn")
        raise HTTPException(status_code=401, detail="Incorrect credential")
    session = secrets.token_urlsafe(24)
    ADMIN_SESSIONS[session] = datetime.now(timezone.utc).timestamp()
    audit("admin_auth_login", detail="Admin session opened")
    return {"status": "ok", "session": session}


@app.post("/admin/auth/logout")
async def admin_auth_logout(x_admin_session: Optional[str] = Header(default=None)):
    if x_admin_session:
        ADMIN_SESSIONS.pop(x_admin_session, None)
    return {"status": "ok"}


@app.get("/admin/config")
async def admin_config(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    return _config_db()


@app.post("/admin/config")
async def admin_config_save(payload: ConfigPayload, x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    tokens = [t.strip().upper() for t in payload.admin_tokens if t.strip()]
    _save_config_db({"admin_tokens": tokens, "updated_at": _now_iso()})
    audit("admin_config_saved", detail=f"Configured admin tokens: {', '.join(tokens) or 'none'}")
    return {"status": "ok", "admin_tokens": tokens}


@app.post("/wizard/progress")
async def wizard_progress_save(payload: WizardRecord):
    token = payload.token.strip().upper()
    if token not in users:
        raise HTTPException(status_code=404, detail="Unknown token")
    db = _wizard_db()
    row = payload.model_dump()
    row["token"] = token
    row["updated_at"] = _now_iso()
    db[token] = row
    _save_wizard_db(db)
    audit("wizard_progress_saved", token=token, detail="Wizard profile/progress updated")
    return {"status": "ok", "record": row}


@app.get("/wizard/progress/{token}")
async def wizard_progress_get(token: str):
    token = token.strip().upper()
    if token not in users:
        raise HTTPException(status_code=404, detail="Unknown token")
    db = _wizard_db()
    return db.get(token, {
        "token": token,
        "display_name": users[token]["name"],
        "iits_username": "",
        "adm_username": "",
        "completed": [],
        "adminCompleted": [],
        "iits_pw_date": None,
        "adm_pw_date": None,
        "vpn_date": None,
    })


@app.get("/admin/wizard")
async def admin_wizard(x_admin_session: Optional[str] = Header(default=None)):
    _require_admin(x_admin_session)
    db = _wizard_db()
    merged = []
    for token, u in sorted(users.items()):
        rec = db.get(token, {})
        merged.append({
            "token": token,
            "display_name": rec.get("display_name") or u.get("name", ""),
            "email": u.get("email", ""),
            "iits_username": rec.get("iits_username", ""),
            "adm_username": rec.get("adm_username", ""),
            "completed": rec.get("completed", []),
            "adminCompleted": rec.get("adminCompleted", []),
            "iits_pw_date": rec.get("iits_pw_date"),
            "adm_pw_date": rec.get("adm_pw_date"),
            "vpn_date": rec.get("vpn_date"),
            "updated_at": rec.get("updated_at"),
        })
    return {"users": merged}


@app.post("/api/onboard/notify")
async def onboard_notify(request: Request):
    payload = await request.json()
    token = str(payload.get("token", "") or "").strip().upper() or None
    detail = json.dumps(payload, sort_keys=True)[:500]
    audit("onboard_notify", token=token, detail=detail)
    return {"status": "ok", "received": payload, "ts": _now_iso()}

# Serve frontend — must be last
app.mount("/", StaticFiles(directory="frontend", html=True), name="frontend")
